from hashlib import sha256
import pandas as pd
import qrcode as qr
import logging
import argparse
import pathlib
from openpyxl import load_workbook

WHITELIST_DIR = pathlib.Path("data/whitelist.xlsx")
QR_DIR = pathlib.Path("data/output_qr")

def generate_hash_key(aa : str, name : str) -> str:
    input_str = f"{aa.strip()}-{name.lower().strip()}"
    return sha256(input_str.encode('utf-8')).hexdigest()

def generate_qr_image(hash: str, output_path: pathlib.Path):    
        qr_image = qr.make(hash)
        qr_image.save(output_path)

def generate_whitelist(entries: list[dict]):
    WHITELIST_DIR.parent.mkdir(exist_ok=True, parents=True)
    QR_DIR.mkdir(exist_ok=True, parents=True)

    df = pd.DataFrame(entries)
         
    if WHITELIST_DIR.exists():
        book = load_workbook(WHITELIST_DIR)
        writer = pd.ExcelWriter(WHITELIST_DIR, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        startrow = writer.sheets['Sheet1'].max_row
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=startrow)
        writer.close()
    else:
        df.to_excel(WHITELIST_DIR, index=False)
    logging.info(f"Appended {len(entries)} entries to {WHITELIST_DIR}")    


def main():
    logging.info(f"Reading Excel source: {args.source}")
    df = pd.read_excel(args.source)
    whitelist_entries = []
    for idx, row in df.iterrows():
        
        status = str(row.get("ΚΑΤΑΣΤΑΣΗ ΔΕΛΤΙΟΥ", "")).strip().upper()
        
        if status != "ΙΣΧΥΕΙ":
            logging.info(f"Skipping row {idx} — ΚΑΤΑΣΤΑΣΗ ΔΕΛΤΙΟΥ: '{status}'")
            continue
        
        num = str(row.get("Α/Α ΔΕΛΤΙΟΥ", "")).strip()
        name = str(row.get("EΠΩΝΥΜΟ", "")).strip()
        
        if not num or not name:
            logging.warning(f"Skipping row {idx} — missing Α/Α ΔΕΛΤΙΟΥ or ΕΠΩΝΥΜΟ")
            continue
        hash = generate_hash_key(num, name)
        qr_filename = f"{num}-{name}.png"
        qr_path = QR_DIR / qr_filename
        generate_qr_image(hash, qr_path)

        entry = {
            "Α/Α ΔΕΛΤΙΟΥ": num,
            "ΕΠΩΝΥΜΟ": name,
            "qr_data": hash,
            "qr_image": str(qr_path)
        }
        whitelist_entries.append(entry)

    generate_whitelist(whitelist_entries)
    logging.info("Finished generating QR codes and creating whitelist")

if __name__ == "__main__":

    log = logging.getLogger()
    log.setLevel(logging.INFO)
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    stream_handler.set_name('stdout')
    formatter = logging.Formatter('[%(levelname)s:%(funcName)s] %(message)s')
    stream_handler.setFormatter(formatter)
    log.addHandler(stream_handler)

    args = argparse.ArgumentParser()
    args.add_argument("--source", action="store", required=True)
    args.add_argument("--ignore_names", action="store", type=pathlib.Path, default="ignore.txt")
    args = args.parse_args()
    
    main()

